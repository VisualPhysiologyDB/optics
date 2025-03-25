# prediction_functions.py
import subprocess 
from deepBreaks.utils import load_obj
from deepBreaks.preprocessing import read_data
import pandas as pd
import json
import argparse
import os
import sys
import numpy as np
import pathlib  # pathlib is generally better than os.path (see below)
import datetime
import matplotlib
from joblib import Parallel, delayed
import tempfile
from multiprocessing import Manager
from tqdm import tqdm
from optics_scripts.blastp_align import seq_sim_report
from optics_scripts.bootstrap_predictions import calculate_ensemble_CI, plot_prediction_subsets_with_CI, wavelength_to_rgb

def extract_fasta(file):
    with open(file, 'r') as f:
        sequences = []
        names = []
        i = 0
        line_count = 0
        entry = ""
        lines = f.readlines()
        num_lines = len(lines)
        #print(num_lines)

        for line in lines:
            if '>' in line:
                if i == 1:
                    names.append(line.replace('>','').strip().replace(' ','_'))
                    sequences.append(entry)
                    entry = ""
                    entry += line
                    #print(sequences)
                    line_count+=1
                else:
                    names.append(line.replace('>','').strip().replace(' ','_'))
                    entry += line
                    i+=1
                    line_count+=1
            else:
                entry += line
                line_count+=1
                if line_count >= num_lines:
                     sequences.append(entry)
    #print(sequences)
    #print(names)
    return names,sequences

def process_sequence(sequence=None, name=None, selected_model=None, identity_report=None, blastp=None, refseq=None, reffile=None, bootstrap=None, prediction_dict=None, encoding_method='one_hot', wrk_dir = '', only_blast = False):
    if wrk_dir == '':
        script_path = pathlib.Path(__file__).resolve()  # Get absolute path
        wrk_dir = str(script_path.parent).replace('\\', '/')

    data_dir = f"{wrk_dir}/data"
    #print(f'This is the data_dir: {data_dir}')
    model_datasets = {
        "whole-dataset": f"{data_dir}/fasta/vpod_1.2/wds_aligned_VPOD_1.2_het.fasta",
        "wildtype": f"{data_dir}/fasta/vpod_1.2/wt_aligned_VPOD_1.2_het.fasta",
        "vertebrate": f"{data_dir}/fasta/vpod_1.2/vert_aligned_VPOD_1.2_het.fasta",
        "invertebrate": f"{data_dir}/fasta/vpod_1.2/inv_only_aligned_VPOD_1.2_het.fasta",
        "wildtype-vert": f"{data_dir}/fasta/vpod_1.2/wt_vert_aligned_VPOD_1.2_het.fasta",
        "type-one": f"{data_dir}/fasta/vpod_1.2/Karyasuyama_T1_ops_aligned.fasta",
        "whole-dataset-mnm": f"{data_dir}/fasta/vpod_1.2/wds_mnm_aligned_VPOD_1.2_het.fasta",
        "wildtype-mnm": f"{data_dir}/fasta/vpod_1.2/wt_mnm_aligned_VPOD_1.2_het.fasta",
        "vertebrate-mnm": f"{data_dir}/fasta/vpod_1.2/vert_mnm_aligned_VPOD_1.2_het.fasta",
        "invertebrate-mnm": f"{data_dir}/fasta/vpod_1.2/inv_mnm_aligned_VPOD_1.2_het.fasta",
        "wildtype-vert-mnm": f"{data_dir}/fasta/vpod_1.2/wt_vert_mnm_aligned_VPOD_1.2_het.fasta",
    }   
    model_raw_data = {
        "whole-dataset": f"{data_dir}/fasta/vpod_1.2/wds.txt",
        "wildtype": f"{data_dir}/fasta/vpod_1.2/wt.txt",
        "vertebrate": f"{data_dir}/fasta/vpod_1.2/vert.txt",
        "invertebrate": f"{data_dir}/fasta/vpod_1.2/inv_only.txt",
        "wildtype-vert": f"{data_dir}/fasta/vpod_1.2/wt_vert.txt",
        "type-one": f"{data_dir}/fasta/vpod_1.2/Karyasuyama_T1_ops.txt",
        "whole-dataset-mnm": f"{data_dir}/fasta/vpod_1.2/wds_mnm.txt",
        "wildtype-mnm": f"{data_dir}/fasta/vpod_1.2/wt_mnm.txt",
        "vertebrate-mnm": f"{data_dir}/fasta/vpod_1.2/vert_mnm.txt",
        "invertebrate-mnm": f"{data_dir}/fasta/vpod_1.2/inv_mnm.txt",
        "wildtype-vert-mnm": f"{data_dir}/fasta/vpod_1.2/wt_vert_mnm.txt",
    }   
    model_metadata = {
        "whole-dataset": f"{data_dir}/fasta/vpod_1.2/wds_meta.tsv",
        "wildtype": f"{data_dir}/fasta/vpod_1.2/wt_meta.tsv",
        "vertebrate": f"{data_dir}/fasta/vpod_1.2/vert_meta.tsv",
        "invertebrate": f"{data_dir}/fasta/vpod_1.2/inv_meta.tsv",
        "wildtype-vert": f"{data_dir}/fasta/vpod_1.2/wt_vert_meta.tsv",
        "type-one": f"{data_dir}/fasta/vpod_1.2/Karyasuyama_T1_ops_meta.tsv",
        "whole-dataset-mnm": f"{data_dir}/fasta/vpod_1.2/wds_mnm_meta.csv",
        "wildtype-mnm": f"{data_dir}/fasta/vpod_1.2/wt_mnm_meta.csv",
        "vertebrate-mnm": f"{data_dir}/fasta/vpod_1.2/vert_mnm_meta.csv",
        "invertebrate-mnm": f"{data_dir}/fasta/vpod_1.2/inv_mnm_meta.csv",
        "wildtype-vert-mnm": f"{data_dir}/fasta/vpod_1.2/wt_vert_mnm_meta.csv",
    }   
    model_blast_db = {
        "whole-dataset": f"{data_dir}/blast_dbs/vpod_1.2/wds_db",
        "wildtype": f"{data_dir}/blast_dbs/vpod_1.2/wt_db",
        "vertebrate": f"{data_dir}/blast_dbs/vpod_1.2/vert_db",
        "invertebrate": f"{data_dir}/blast_dbs/vpod_1.2/invert_db",
        "wildtype-vert": f"{data_dir}/blast_dbs/vpod_1.2/wt_vert_db",
        "type-one": f"{data_dir}/blast_dbs/vpod_1.2/t1_db",
        "whole-dataset-mnm": f"{data_dir}/blast_dbs/vpod_1.2/wds_mnm_db",
        "wildtype-mnm": f"{data_dir}/blast_dbs/vpod_1.2/wt_mnm_db",
        "vertebrate-mnm": f"{data_dir}/blast_dbs/vpod_1.2/vert_mnm_db",
        "invertebrate-mnm": f"{data_dir}/blast_dbs/vpod_1.2/inv_mnm_db",
        "wildtype-vert-mnm": f"{data_dir}/blast_dbs/vpod_1.2/wt_vert_mnm_db",
    }   

    model_dir = f"{wrk_dir}/models"

    if encoding_method == 'aa_prop':
        model_directories = {
            "whole-dataset": f"{model_dir}/reg_models/vpod_1.2/aa_prop/wds_xgb.pkl",
            "wildtype": f"{model_dir}/reg_models/vpod_1.2/aa_prop/wt_gbr.pkl",
            "vertebrate": f"{model_dir}/reg_models/vpod_1.2/aa_prop/vert_xgb.pkl",
            "invertebrate": f"{model_dir}/reg_models/vpod_1.2/aa_prop/inv_gbr.pkl",
            "wildtype-vert": f"{model_dir}/reg_models/vpod_1.2/aa_prop/wt_vert_gbr.pkl",
            "type-one": f"{model_dir}/reg_models/vpod_1.2/aa_prop/t1_xgb.pkl",
            "whole-dataset-mnm": f"{model_dir}/reg_models/vpod_1.2/aa_prop/wds_mnm_xgb.pkl",
            "wildtype-mnm": f"{model_dir}/reg_models/vpod_1.2/aa_prop/wt_mnm_gbr.pkl",
            "vertebrate-mnm": f"{model_dir}/reg_models/vpod_1.2/aa_prop/vert_mnm_xgb.pkl",
            "invertebrate-mnm": f"{model_dir}/reg_models/vpod_1.2/aa_prop/invert_mnm_gbr.pkl",
            "wildtype-vert-mnm": f"{model_dir}/reg_models/vpod_1.2/aa_prop/wt_vert_mnm_xgb.pkl",
        }
        
        model_bs_dirs = {
            "whole-dataset": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/wds_H1_H2_H3_P2_V_SCT_PKA_bootstrap_100_2025-03-23_22-05-50",
            "wildtype": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/wt_H2_H3_P1_NCI_PKA_bootstrap_100_2025-03-24_10-19-45",
            "vertebrate": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/vert_H2_H3_NCI_SCT_PKB_bootstrap_100_2025-03-21_17-47-40",
            "invertebrate": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/inv_H1_H3_bootstrap_100_2025-03-21_17-40-08",
            "wildtype-vert": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/wt_vert_H2_P2_V_MASS_bootstrap_100_2025-03-21_17-25-47",
            "type-one": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/t1_H3_P1_PKB_bootstrap_100_2025-03-24_10-31-03",
            "whole-dataset-mnm": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/wds_mnm_bootstrap",
            "wildtype-mnm": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/wt_mnm_bootstrap",
            "vertebrate-mnm": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/vert_mnm_bootstrap",
            "invertebrate-mnm": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/invert_mnm_bootstrap",
            "wildtype-vert-mnm": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/wt_vert_mnm_bootstrap",
        }
        
    else:
        model_directories = {
            "whole-dataset": f"{model_dir}/reg_models/vpod_1.2/one_hot/wds_xgb.pkl",
            "wildtype": f"{model_dir}/reg_models/vpod_1.2/one_hot/wt_xgb.pkl",
            "vertebrate": f"{model_dir}/reg_models/vpod_1.2/one_hot/vert_xgb.pkl",
            "invertebrate": f"{model_dir}/reg_models/vpod_1.2/one_hot/invert_BayesianRidge.pkl",
            "wildtype-vert": f"{model_dir}/reg_models/vpod_1.2/one_hot/wt_vert_xgb.pkl",
            "type-one": f"{model_dir}/reg_models/vpod_1.2/one_hot/t1_xgb.pkl",
            "whole-dataset-mnm": f"{model_dir}/reg_models/vpod_1.2/one_hot/wds_mnm_xgb.pkl",
            "wildtype-mnm": f"{model_dir}/reg_models/vpod_1.2/one_hot/wt_mnm_gbr.pkl",
            "vertebrate-mnm": f"{model_dir}/reg_models/vpod_1.2/one_hot/vert_mnm_xgb.pkl",
            "invertebrate-mnm": f"{model_dir}/reg_models/vpod_1.2/one_hot/invert_mnm_gbr.pkl",
            "wildtype-vert-mnm": f"{model_dir}/reg_models/vpod_1.2/one_hot/wt_vert_mnm_xgb.pkl",
        }
        
        model_bs_dirs = {
            "whole-dataset": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/wds_bootstrap",
            "wildtype": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/wt_bootstrap",
            "vertebrate": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/vert_bootstrap",
            "invertebrate": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/invert_bootstrap",
            "wildtype-vert": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/wt_vert_bootstrap",
            "type-one": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/t1_bootstrap",
            "whole-dataset-mnm": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/wds_mnm_bootstrap",
            "wildtype-mnm": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/wt_mnm_bootstrap",
            "vertebrate-mnm": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/vert_mnm_bootstrap",
            "invertebrate-mnm": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/invert_mnm_bootstrap",
            "wildtype-vert-mnm": f"{model_dir}/bs_models/vpod_1.2/{encoding_method}/wt_vert_mnm_bootstrap",
        }
        

    if sequence == None:
        return ('Error: No sequence given')
    #print(sequence)
    if selected_model == None:
        return ('Error: No model selected')
    
    alignment_data = model_datasets[selected_model]
    raw_data = model_raw_data[selected_model]
    metadata = model_metadata[selected_model]
    blast_db = model_blast_db[selected_model]
    model_bs_folder = model_bs_dirs[selected_model]
    model = model_directories[selected_model]
    
    #wrk_dir = os.getcwd().replace('\\','/')
    with tempfile.NamedTemporaryFile(mode="w", dir=f"{wrk_dir}/tmp", suffix=".fasta", delete=False) as temp_seq_file:
        temp_seq = temp_seq_file.name  # Get the unique filename
        #print(temp_seq)
        if '>' in sequence:
            temp_seq_file.write(sequence)
        else:
            sequence = ">placeholder_name\n" + sequence
            temp_seq_file.write(sequence)
    
    # This is a special case for when a cached sequence is detected
    if only_blast == True:
        if blastp == 'no' or blastp == False or blastp == 'False':
            percent_iden = '-'
        else:
            percent_iden = seq_sim_report(temp_seq, name, refseq, blast_db, raw_data, metadata, identity_report, reffile, wrk_dir)
        
        os.remove(temp_seq)
        return percent_iden

    if blastp == 'no' or blastp == False or blastp == 'False':
        percent_iden = '-'
    else:
        percent_iden = seq_sim_report(temp_seq, name, refseq, blast_db, raw_data, metadata, identity_report, reffile, wrk_dir)
        #print('Query sequence processed via blastp')

    with tempfile.NamedTemporaryFile(mode="w", dir=f"{wrk_dir}/tmp", suffix=".fasta", delete=False) as temp_ali_file:
        new_ali = temp_ali_file.name 
    try:
        #print('Trying Linux execution of MAFFT')
        cmd = ['mafft', '--add', temp_seq, '--keeplength', alignment_data]
        with open(new_ali, 'w') as f:
            subprocess.run(cmd, stdout=f, stderr=subprocess.PIPE)
    except:
        #print('Trying Windows execution of MAFFT')
        try:
            mafft_exe = f'{wrk_dir}/optics_scripts/mafft/mafft-win/mafft.bat'
            cmd = [mafft_exe, '--add', temp_seq, '--keeplength', alignment_data]
            with open(new_ali, 'w') as f:
                subprocess.run(cmd, stdout=f, stderr=subprocess.PIPE)
        except: 
            #print('Trying Mac execution of MAFFT')
            try:
                mafft_exe = f'{wrk_dir}/optics_scripts/mafft/mafft-mac/mafft.bat'
                cmd = [mafft_exe, '--add', temp_seq, '--keeplength', alignment_data]
                with open(new_ali, 'w') as f:
                    subprocess.run(cmd, stdout=f, stderr=subprocess.PIPE)
            except subprocess.CalledProcessError as e:
                raise Exception(f'MAFFT alignment failed for all three options (Linux, Windows, and Max).\nCheck your FASTA file to make sure it is formatted correctly as that could be a source of error.\n{e.stderr.decode()}')
            
    seq_type = 'aa'
    
    new_seq_test = read_data(new_ali, seq_type = seq_type, is_main=True, gap_threshold=0.5)
    ref_copy = read_data(alignment_data, seq_type = seq_type, is_main=True, gap_threshold=0.5)
    last_seq = int(ref_copy.shape[0])
    new_seq_test = new_seq_test.iloc[last_seq:].copy()
    #print(new_seq_test)

    try:
        os.remove(new_ali)
        os.remove(temp_seq)
    except FileNotFoundError:
        raise Exception("File does not exist")

    if bootstrap == True:
        # ... (Load the selected model and make a bootstrap prediction)
        mean_prediction, ci_lower, ci_upper, prediction_dict, median_prediction, std_dev, predictions_all = calculate_ensemble_CI(model_bs_folder, new_seq_test, name, prediction_dict)
        #print(f'{mean_prediction}, {ci_lower}, {ci_upper}, {prediction_dict}')
        
        # ... (Load the selected model and make a single prediction)
        loaded_mod = load_obj(model)
        prediction = loaded_mod.predict(new_seq_test)
        
        return (round(float(mean_prediction),1), round(float(ci_lower),1), round(float(ci_upper),1), prediction_dict, round(float(prediction[0]),1), round(float(median_prediction),1), str(percent_iden), round(float(std_dev),1), predictions_all.tolist())

    else:
        # ... (Load the selected model and make a prediction)
        loaded_mod = load_obj(model)
        prediction = loaded_mod.predict(new_seq_test)
        
        return(round(float(prediction[0]),1), str(percent_iden))
 

def process_sequences_from_file(file, selected_model, identity_report, blastp, refseq, reffile, bootstrap, encoding_method, wrk_dir):

    cache_dir = f"{wrk_dir}/data/cached_predictions"
    if (bootstrap == True or bootstrap == 'True' or bootstrap == 'true' or bootstrap == 'yes'):
        bootstrap = True
        model_type = 'bs_models'
    else:
        bootstrap = False
        model_type = 'reg_models'
    model_cache_dirs = {
        "whole-dataset": f"{cache_dir}/{model_type}/vpod_1.2/{encoding_method}/wds_pred_dict.json",
        "wildtype": f"{cache_dir}/{model_type}/vpod_1.2/{encoding_method}/wt_pred_dict.json",
        "vertebrate": f"{cache_dir}/{model_type}/vpod_1.2/{encoding_method}/vert_pred_dict.json",
        "invertebrate": f"{cache_dir}/{model_type}/vpod_1.2/{encoding_method}/invert_pred_dict.json",
        "wildtype-vert": f"{cache_dir}/{model_type}/vpod_1.2/{encoding_method}/wt_vert_pred_dict.json",
        "type-one": f"{cache_dir}/{model_type}/vpod_1.2/{encoding_method}/t1_pred_dict.json",
        "whole-dataset-mnm": f"{cache_dir}/{model_type}/vpod_1.2/{encoding_method}/wds_mnm_pred_dict.json",
        "wildtype-mnm": f"{cache_dir}/{model_type}/vpod_1.2/{encoding_method}/wt_mnm_pred_dict.json",
        "vertebrate-mnm": f"{cache_dir}/{model_type}/vpod_1.2/{encoding_method}/vert_mnm_pred_dict.json",
        "invertebrate-mnm": f"{cache_dir}/{model_type}/vpod_1.2/{encoding_method}/invert_mnm_pred_dict.json",
        "wildtype-vert-mnm": f"{cache_dir}/{model_type}/vpod_1.2/{encoding_method}/wt_vert_mnm_pred_dict.json",
    }

    cache_file = model_cache_dirs[selected_model]
    
    if file == None:
        raise Exception('Error: No file given')
    
    names,sequences = extract_fasta(file)       
    predictions = []
    mean_predictions = []
    median_predictions = []
    std_dev_list = []
    ci_lowers = []
    ci_uppers = []
    per_iden_list = []
    seq_lens = []
    # Load the existing prediction dictionary to pull from if a sequence has already been predicted by the chosen model
    # Check to see if an existing prediiction dictionary already exists to save time.
    # In this dictionary the sequence will be our keys, and the len(seq), prediction, percent_iden, mean_prediction, ci_lower, ci_upper, median_prediction, std_dev will be the corresponding values...
    if os.path.isfile(cache_file):
        try:
            with open(cache_file, 'r') as f:
                cached_pred_dict = json.load(f)
        except FileNotFoundError:
            cached_pred_dict = {}
    else:
        cached_pred_dict = {}
    
    manager = Manager()
    prediction_dict = manager.dict()  # Use a shared dictionary for prediction outputs
    mp_cached_pred_dict = manager.dict(cached_pred_dict)  # Initialize a multiprocess available cached dict with cached data
    
    def process_sequence_wrapper(seq, name):  # Helper function
        just_seq = seq.split('\n')[1]
        if bootstrap == False:
            if just_seq not in list(mp_cached_pred_dict.keys()):
                prediction, percent_iden = process_sequence(seq, name, selected_model, identity_report, blastp, refseq, reffile, bootstrap, prediction_dict, encoding_method, wrk_dir)
                mp_cached_pred_dict[just_seq] = {'len': len(just_seq), 'single_prediction': prediction}
                return len(just_seq), prediction, percent_iden, None, None, None, None, None  # Consistent return values
            else:
                percent_iden = process_sequence(seq, name, selected_model, identity_report, blastp, refseq, reffile, bootstrap, prediction_dict, encoding_method, wrk_dir, only_blast=True)
                return mp_cached_pred_dict[just_seq]['len'], mp_cached_pred_dict[just_seq]['single_prediction'], percent_iden, None, None, None, None, None  # Consistent return values

        else:
            if just_seq not in list(mp_cached_pred_dict.keys()):
                mean_prediction, ci_lower, ci_upper, updated_prediction_dict, prediction, median_prediction, percent_iden, std_dev, predictions_all = process_sequence(seq, name, selected_model, identity_report, blastp, refseq, reffile, bootstrap, prediction_dict, encoding_method, wrk_dir)
                prediction_dict.update(updated_prediction_dict) # Update the shared dictionary with the returned dictionary
                mp_cached_pred_dict[just_seq] = {'len': len(just_seq),
                                            'single_prediction': prediction,
                                            'mean_prediction' : mean_prediction,
                                            'ci_lower' : ci_lower,
                                            'ci_upper' : ci_upper,
                                            'median_prediction' : median_prediction,
                                            'std_deviation' : std_dev,
                                            'all_bs_predictions' : predictions_all
                                            }
                # If the seq is in our cache, I need to be able to update the prediction_dict with that info...
                return len(just_seq), prediction, percent_iden, mean_prediction, ci_lower, ci_upper, median_prediction, std_dev
            else:
                percent_iden = process_sequence(seq, name, selected_model, identity_report, blastp, refseq, reffile, bootstrap, prediction_dict, encoding_method, wrk_dir, only_blast=True)
                prediction_dict[name] = np.array(mp_cached_pred_dict[just_seq]['all_bs_predictions'])
                return mp_cached_pred_dict[just_seq]['len'], mp_cached_pred_dict[just_seq]['single_prediction'], percent_iden, mp_cached_pred_dict[just_seq]['mean_prediction'], mp_cached_pred_dict[just_seq]['ci_lower'], mp_cached_pred_dict[just_seq]['ci_upper'], mp_cached_pred_dict[just_seq]['median_prediction'], mp_cached_pred_dict[just_seq]['std_deviation']


    try:
        with tqdm_joblib(tqdm(total=len(sequences), desc="Processing Sequences", bar_format="{l_bar}{bar:25}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}{postfix}]", 
                              dynamic_ncols=True, colour="#CF9FFF",
                              unit ='seqs',ascii="░▒▓")) as pbar:  # Use tqdm for progress bar
            results = Parallel(n_jobs=-1)(delayed(process_sequence_wrapper)(seq, names[i]) for i, seq in enumerate(sequences))
    except Exception as e:
        updated_cached_pred_dict = dict(mp_cached_pred_dict)        
        # Save the taxon dictionary if it doesn't yet exist or if it has been updated since being loaded 
        if list(updated_cached_pred_dict.keys()) > list(cached_pred_dict.keys()):  
            #print('Saving Updated Dictionary') 
            try:
                with open(cache_file, 'w') as f:
                    json.dump(updated_cached_pred_dict, f, indent=4)  # indent for pretty formatting
                    raise Exception(f"Error Occured During the Prediction Process:\n{e}\n")

            except FileNotFoundError:
                raise Exception(f"Error: Cached prediction file can't be saved...\n")
        
        raise Exception(f"Error: {e}...\n")


    # Extract results
    for result in results:
        seq_len, prediction, percent_iden, mean_prediction, ci_lower, ci_upper, median_prediction, std_dev = result
        seq_lens.append(seq_len)
        predictions.append(prediction)
        per_iden_list.append(percent_iden)
        if mean_prediction is not None:  # Handle bootstrap case
            mean_predictions.append(mean_prediction)
            ci_lowers.append(ci_lower)
            ci_uppers.append(ci_upper)
            median_predictions.append(median_prediction)
            std_dev_list.append(std_dev)
            
    updated_cached_pred_dict = dict(mp_cached_pred_dict)        
    # Save the taxon dictionary if it doesn't yet exist or if it has been updated since being loaded 
    if list(updated_cached_pred_dict.keys()) > list(cached_pred_dict.keys()):  
        #print('Saving Updated Dictionary') 
        try:
            with open(cache_file, 'w') as f:
                json.dump(updated_cached_pred_dict, f, indent=4)  # indent for pretty formatting
        except FileNotFoundError:
            print(f"Error: Cached prediction file can't be saved...\n")

    return(names, mean_predictions, ci_lowers, ci_uppers, prediction_dict, predictions, median_predictions, per_iden_list, std_dev_list, seq_lens)

def run_optics_predictions(input_sequence, pred_dir=None, output='optics_predictions.txt',
                           model="whole-dataset", encoding_method='aa_prop', blastp=True,
                           iden_report='blastp_report.txt', refseq='bovine', reffile=None,
                           bootstrap=True, visualize_bootstrap=True, bootstrap_viz_file='bootstrap_viz'):
    """
    Processes sequences using a selected model and generates prediction outputs.  This function
    encapsulates the logic from the original `main` function, making it callable from
    other scripts or notebooks.

    Args:
        input_sequence (str): Either a single sequence or a path to a FASTA file.
        report_dir (str, optional):  Name of folder directory to create.  Defaults to a
            timestamped directory if None.  If a directory name is provided, a timestamp
            will be appended to avoid overwrites, and all results put in a 'prediction_outputs' subdirectory.
        output (str, optional): Name for output file. Defaults to 'optics_predictions.txt'.
        model (str, optional): Model to use for prediction. Defaults to "whole-dataset".
        encoding_method (str, optional): Encoding method. Defaults to 'aa_prop'.
        blastp (bool, optional): Enable blastp analysis. Defaults to True.
        iden_report (str, optional): Blastp report output file name. Defaults to 'blastp_report.txt'.
        refseq (str, optional): Reference sequence for blastp. Defaults to 'bovine'.
        reffile (str, optional): Custom reference sequence file. Defaults to None.
        bootstrap (bool, optional): Enable bootstrap predictions. Defaults to True.
        visualize_bootstrap (bool, optional): Enable visualization of bootstrap predictions. Defaults to True.
        bootstrap_viz_file (str, optional): Output file name for bootstrap visualization. Defaults to 'bootstrap_viz'.

    Returns:
        tuple: A tuple containing the following lists (in order):
            - names: List of sequence names.
            - mean_predictions: List of mean predictions (if bootstrapping).
            - ci_lowers: List of lower confidence interval bounds (if bootstrapping).
            - ci_uppers: List of upper confidence interval bounds (if bootstrapping).
            - prediction_dict: Dictionary of all bootstrap predictions (if bootstrapping).
            - predictions: List of single predictions (or first prediction from bootstrap).
            - median_predictions: List of median predictions (if bootstrapping).
            - per_iden_list: List of percent identities from BLASTp.
            - std_dev_list: List of standard deviations (if bootstrapping).
            - seq_lens_list: list of sequence lengths

        Also creates several output files within the specified report directory:
            - Main results file (TSV or TXT)
            - Excel file with additional formatting
            - BLASTp report (if blastp is enabled)
            - Bootstrap visualization (PDF, if bootstrapping and visualization are enabled)
            - Argument log file
            - Color annotation files for FigTree and iTOL

    Raises:
        Exception: If `input_sequence` is not a file or a valid sequence.  (Improved error handling.)
    """
    dt_label = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    script_path = pathlib.Path(__file__).resolve()  # Get absolute path
    wrk_dir = str(script_path.parent).replace('\\', '/')
    #print(f"Script directory (pathlib): {wrk_dir}")

    # Argument validation (mimicking argparse behavior, but for function inputs)
    models = ['whole-dataset', 'wildtype', 'vertebrate', 'invertebrate', 'wildtype-vert', 'type-one', 'whole-dataset-mnm', 'wildtype-mnm', 'vertebrate-mnm', 'invertebrate-mnm', 'wildtype-vert-mnm']
    encoding_methods = ['one_hot', 'aa_prop']
    ref_seq_choices = ['bovine', 'squid', 'microbe', 'custom']
    #bool_choices = ['true', 'True', 'false', 'False', True, False] #not used, we cast directly to bool.

    if model not in models:
        raise ValueError(f"Invalid model choice.  Must be one of {models}")
    if encoding_method not in encoding_methods:
        raise ValueError(f"Invalid encoding method choice.  Must be one of {encoding_methods}")
    if refseq not in ref_seq_choices:
        raise ValueError(f"Invalid refseq choice. Must be one of {ref_seq_choices}")
    # No need to check bool_choices.  We'll convert directly to boolean.

    # Directory setup (with added handling for pre-existing directories)
    if not os.path.isdir(f'{wrk_dir}/tmp'):
        os.makedirs(f'{wrk_dir}/tmp')
    if not os.path.isdir('./prediction_outputs'):
        os.makedirs('./prediction_outputs')

    if pred_dir is None:
        report_dir = f'./prediction_outputs/optics_on_unamed_{dt_label}'
    else:
        report_dir = f'./prediction_outputs/optics_on_{pred_dir}_{dt_label}'
    os.makedirs(report_dir, exist_ok=True)  # exist_ok=True prevents errors if dir exists



    blastp_file = f'{report_dir}/{iden_report}'
    if not (blastp_file.endswith('.txt') or blastp_file.endswith('.tsv')):
           blastp_file += '.txt'

    bootstrap_file = f'{report_dir}/{bootstrap_viz_file}'
    log_file = f'{report_dir}/arg_log.txt'

    # Input handling (file or sequence string)
    if os.path.isfile(input_sequence):
        names, mean_predictions, ci_lowers, ci_uppers, prediction_dict, predictions, median_predictions, per_iden_list, std_dev_list, seq_lens_list = process_sequences_from_file(input_sequence, model, blastp_file, blastp, refseq, reffile, bootstrap, encoding_method, wrk_dir)
        
        # Output file handling (TSV or TXT, Excel)
        if output.endswith(('.tsv', '.txt')):
            output_path = f'{report_dir}/{output}'
            excel_output = f'{report_dir}/{output.replace(".tsv", "").replace(".txt", "")}_for_excel.xlsx'
        else:
            output_path = f'{report_dir}/{output}.tsv'
            excel_output = f'{report_dir}/{output}_for_excel.xlsx'

    else:  # Assume it's a single sequence
        #  create a temporary file.
        temp_input_file = os.path.join('./tmp', f'temp_input_{dt_label}.fasta')
        with open(temp_input_file, "w") as f:
            f.write(f">temp_seq\n{input_sequence}\n") #write temp file to be consistant with process_sequence_from_file function
        
        names, mean_predictions, ci_lowers, ci_uppers, prediction_dict, predictions, median_predictions, per_iden_list, std_dev_list, seq_lens_list = process_sequences_from_file(temp_input_file, model, blastp_file, blastp, refseq, reffile, bootstrap, encoding_method, wrk_dir)
        output_path = f'{report_dir}/{output}'
        if not output_path.endswith(('.tsv', '.txt')):
            output_path += '.tsv'
        excel_output = output_path.replace('.tsv', '_for_excel.xlsx').replace('.txt', '_for_excel.xlsx')
        os.remove(temp_input_file) #clean up temp file


    # Write main results file
    with open(output_path, 'w') as f:
        if not bootstrap: #cast to bool to simplify logic
            colors = [wavelength_to_rgb(pred) for pred in predictions]
            hex_color_list = [matplotlib.colors.to_hex(color) for color in colors]
            write_to_excel(names, predictions, per_iden_list, excel_output, hex_color_list=hex_color_list, seq_lens_list=seq_lens_list)
            # Make Prediction Dataframe
            pred_df = pd.DataFrame({
            'Names': names,
            'Single_Prediction': predictions,
            '%Identity_Nearest_VPOD_Sequence': per_iden_list,
            'Sequence_Length': seq_lens_list,
            'Lmax_Hex_Color': hex_color_list
            })
            # Write to file
            pred_df.to_csv(f, sep='\t', index=False)
            for i in range(len(names)):
                print(f"{names[i]}\t{predictions[i]}\t{per_iden_list[i]}\t{seq_lens_list[i]}\n")
        else:
            hex_color_list = plot_prediction_subsets_with_CI(names, prediction_dict, mean_predictions,
                                                            bootstrap_file, visualize_bootstrap)
            write_to_excel(names, predictions, per_iden_list, excel_output,
                            mean_predictions, median_predictions, ci_lowers,
                            ci_uppers, std_dev_list, hex_color_list, seq_lens_list)
            # Make Prediction Dataframe
            pred_df = pd.DataFrame({
            'Names': names,
            'Single_Prediction': predictions,
            'Prediction_Means': mean_predictions,
            'Prediction_Medians': median_predictions,
            'Prediction_Lower_Bounds': ci_lowers,
            'Prediction_Upper_Bounds': ci_uppers,
            'Std_Deviation': std_dev_list,
            '%Identity_Nearest_VPOD_Sequence': per_iden_list,
            'Sequence_Length': seq_lens_list,
            'Lmax_Hex_Color': hex_color_list
            })
            # Write to file
            pred_df.to_csv(f, sep='\t', index=False)
            
            print('Names\tSingle_Prediction\tPrediction_Means\tPrediction_Medians\tPrediction_Lower_Bounds\tPrediction_Upper_Bounds\tStd_Deviation\t%Identity_Nearest_VPOD_Sequence\tSequence_Length\n')
            for i in range(len(names)):
                print(f"{names[i]}\t{predictions[i]}\t{mean_predictions[i]}\t{median_predictions[i]}\t{ci_lowers[i]}\t{ci_uppers[i]}\t{std_dev_list[i]}\t{per_iden_list[i]}\t{seq_lens_list[i]}\n")


    # Write log file
    with open(log_file, 'w') as f:
        #  We don't have sys.argv, so we reconstruct the equivalent.
        arg_string = (f"input_fasta: {input_sequence}\nreport_dir: {report_dir}\n"
                      f"output_file: {output}\nmodel: {model}\nencoding_method: {encoding_method}\n"
                      f"blastp: {blastp}\nblastp_report: {iden_report}\nrefseq: {refseq}\n"
                      f"custom_ref_file: {reffile}\nbootstrap: {bootstrap}\n"
                      f"visualize_bootstrap: {visualize_bootstrap}\nbootstrap_viz_file: {bootstrap_viz_file}\n")
        
        if blastp == True:
            bp_cmd = f'--blastp --blastp_report {iden_report} --refseq {refseq} '
        else:
            bp_cmd = ''
        if reffile is not None and refseq == 'custom':
            rf_cmd = f"--custom_ref_file {reffile} "
        else:
            rf_cmd = ''
            
        if bootstrap == True:
            bs_cmd = '--bootstrap '
        else:
            bs_cmd = ''
        if visualize_bootstrap == True:
            vb_cmd = f'--visualize_bootstrap --bootstrap_viz_file {bootstrap_viz_file}'
        else:
            vb_cmd = ''
    
        exec_cmd =  (f"python optics_predictions.py " 
                      f"-i {input_sequence} " 
                      f"-o {pred_dir} " 
                      f"-p {output} " 
                      f"-m {model} " 
                      f"-e {encoding_method} " 
                      f"{bp_cmd}" 
                      f"{rf_cmd}" 
                      f"{bs_cmd}" 
                      f"{vb_cmd}\n")
        f.write(f"Selected Options...\n{arg_string}\n")  # More informative
        f.write(f"Command executed (reconstructed): {exec_cmd}\n")
        #f.write(f"Model Used:\t{model}\nEncoding Method:\t{encoding_method}\n")
        print(f"\nModel Used:\t{model}\nEncoding Method:\t{encoding_method}\n")

    # Write color annotation files
    with open(f'{report_dir}/fig_tree_color_annotation.txt', 'w') as g:
        g.write("Name\t!color\n")
        for name, hex_color in zip(names, hex_color_list):
            g.write(f"{name}\t{hex_color}\n")

    with open(f'{report_dir}/itol_color_annotation.txt', 'w') as g:
        g.write("TREE_COLORS\nSEPARATOR TAB\nDATA\n")
        for name, hex_color in zip(names, hex_color_list):
            g.write(f"{name}\tlabel_background\t{hex_color}\n")

    print('Predictions Complete!')

    return pred_df, output_path

    
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def write_to_excel(names, predictions, per_iden_list, output_filename="output.xlsx", 
                  mean_predictions=None, median_predictions=None, ci_lowers=None, 
                  ci_uppers=None, std_dev_list=None, hex_color_list=None, seq_lens_list=None):
    """
    Writes data to an Excel sheet, including bootstrap statistics and
    hexadecimal color codes, and colors the cells based on the hex codes.

    Args:
        names: List of names.
        predictions: List of predictions.
        per_iden_list: List of percentage identities.
        output_filename: Name of the output Excel file.
        mean_predictions: List of mean predictions (optional, for bootstrap).
        median_predictions: List of median predictions (optional, for bootstrap).
        ci_lowers: List of lower confidence intervals (optional, for bootstrap).
        ci_uppers: List of upper confidence intervals (optional, for bootstrap).
        std_dev_list: List of standard deviations (optional, for bootstrap).
        hex_color_list: List of hexadecimal color codes.
        seq_lens_list: List of sequence lengths
        """

    wb = Workbook()
    ws = wb.active

    
    if mean_predictions == None:
        ws.append(['Names', 'Single_Prediction', '%Identity_Nearest_VPOD_Sequence', 'Sequence_Length','Lmax_Hex_Color'])
        for i in range(len(names)):
            # Because openpyxel is picky about hex-codes we need to remove the '#' symbol for it to accept it as a fill color.
            hex_color = hex_color_list[i].replace('#','') 
            ws.append([names[i], predictions[i], per_iden_list[i], seq_lens_list[i], hex_color_list[i]])
            ws.cell(row=i+2, column=5).fill = PatternFill(start_color=hex_color, 
                                                        end_color=hex_color, 
                                                        fill_type="solid")
    else:
        ws.append(['Names', 'Single_Prediction', 'Prediction_Means', 'Prediction_Medians',
                    'Prediction_Lower_Bounds', 'Prediction_Upper_Bounds', 'Std_Deviation', 
                    '%Identity_Nearest_VPOD_Sequence', 'Sequence_Lengths','Lmax_Hex_Color'])
        for i in range(len(names)):
            # Because openpyxel is picky about hex-codes we need to remove the '#' symbol for it to accept it as a fill color.
            hex_color = hex_color_list[i].replace('#','') 
            ws.append([names[i], predictions[i], mean_predictions[i], median_predictions[i],
                        ci_lowers[i], ci_uppers[i], std_dev_list[i], per_iden_list[i], seq_lens_list[i], hex_color_list[i]])
            ws.cell(row=i+2, column=10).fill = PatternFill(start_color=hex_color, 
                                                        end_color=hex_color, 
                                                        fill_type="solid")
    wb.save(output_filename)
    
    
import contextlib  
import joblib  
@contextlib.contextmanager
def tqdm_joblib(tqdm_object):
    """Context manager to patch joblib to report into tqdm progress bar given as argument"""
    class TqdmBatchCompletionCallback(joblib.parallel.BatchCompletionCallBack):
        def __call__(self, *args, **kwargs):
            tqdm_object.update(n=self.batch_size)
            return super().__call__(*args, **kwargs)

    old_batch_callback = joblib.parallel.BatchCompletionCallBack
    joblib.parallel.BatchCompletionCallBack = TqdmBatchCompletionCallback
    try:
        yield tqdm_object
    finally:
        joblib.parallel.BatchCompletionCallBack = old_batch_callback
        tqdm_object.close()
        
if __name__ == '__main__':
    # This part will only execute when the script is run directly (not imported)
    # It provides a simple way to test the module functionality from the command line
        
    parser = argparse.ArgumentParser(description="Predict protein properties using OPTICS.")

    # Input sequence or FASTA file
    parser.add_argument("-i", "--input", 
                        help="Either a single sequence or a path to a FASTA file", 
                        type=str, 
                        required=True)

    # Output directory for all results
    parser.add_argument("-o", "--output_dir", 
                        help="Directory to save output files (optional).", 
                        type=str, 
                        default=".",  # Current directory if not specified
                        required=False)

    # Base filename for predictions
    parser.add_argument("-p", "--prediction_prefix", 
                        help="Base filename for prediction output (optional).", 
                        type=str, 
                        default="optics_predictions", 
                        required=False)

    # Prediction model
    parser.add_argument("-m", "--model", 
                        help="Prediction model to use (optional).", 
                        type=str, 
                        default="whole-dataset", 
                        choices=['whole-dataset', 'wildtype', 'vertebrate', 'invertebrate', 'wildtype-vert', 'type-one', 'whole-dataset-mnm', 'wildtype-mnm', 'vertebrate-mnm', 'invertebrate-mnm', 'wildtype-vert-mnm'],
                        required=False)

    # Encoding method
    parser.add_argument("-e", "--encoding", 
                        help="Encoding method to use (optional).", 
                        type=str, 
                        default="aa_prop",
                        choices=['one_hot', 'aa_prop'],
                        required=False)

    # BLASTp options
    blastp_group = parser.add_argument_group("BLASTp analysis (optional)") # Group related args

    blastp_group.add_argument("--blastp", 
                            help="Enable BLASTp analysis.", 
                            action="store_true")
    blastp_group.add_argument("--blastp_report", 
                            help="Filename for BLASTp report.", 
                            type=str, 
                            default="blastp_report.txt")
    blastp_group.add_argument("--refseq",
                            help="Reference sequence used for blastp analysis.", 
                            type=str, 
                            default="bovine",
                            choices=['bovine', 'squid', 'microbe', 'custom'])
    blastp_group.add_argument("--custom_ref_file",
                            help="Path to a custom reference sequence file for BLASTp.", 
                            type=str)  # No default, as it's optional

    # Bootstrap options
    bootstrap_group = parser.add_argument_group("Bootstrap analysis (optional)")

    bootstrap_group.add_argument("--bootstrap", 
                                help="Enable bootstrap predictions.", 
                                action="store_true")
    bootstrap_group.add_argument("--visualize_bootstrap", 
                                help="Enable visualization of bootstrap predictions.", 
                                action="store_true")
    bootstrap_group.add_argument("--bootstrap_viz_file", 
                                help="Filename prefix for bootstrap visualization (PDF and SVG).", 
                                type=str, 
                                default="bootstrap_viz")
    
    args = parser.parse_args()

    prediction_output = f"{args.prediction_prefix}"
    blastp_report_output = f"{args.blastp_report}" if args.blastp else None
    bootstrap_viz_output = f"{args.bootstrap_viz_file}" if args.visualize_bootstrap else None
    run_optics_predictions(args.input, args.output_dir, prediction_output, args.model, args.encoding,
                        args.blastp, blastp_report_output, args.refseq, args.custom_ref_file,
                        args.bootstrap, args.visualize_bootstrap, bootstrap_viz_output)