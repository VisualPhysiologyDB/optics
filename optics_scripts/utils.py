def extract_fasta_entries(file):
    with open(file, 'r') as f:
        sequences = []
        names = []
        first_entry = True
        line_count = 0
        entry = ""
        lines = f.readlines()
        num_lines = len(lines)
        #print(num_lines)

        for line in lines:
            if '>' in line:
                if first_entry == False:
                    # Append completed entry
                    sequences.append(entry)
                    # Append name of new entry
                    names.append(line.replace('>','').strip().replace(' ','_').replace('\n',''))
                    # Restart with new entry
                    entry = ""
                    entry += line
                    #print(sequences)
                    line_count+=1
                else:
                    # First entry - must declare entry outside the loop, so this is just a neccessary artifact
                    names.append(line.replace('>','').strip().replace(' ','_').replace('\n',''))
                    #print(names)
                    entry += line
                    first_entry = False
                    line_count+=1
            else:
                # This should be adding the new lines of aa data 
                entry += line.strip().replace('\n','')
                line_count+=1
                if line_count >= num_lines:
                     sequences.append(entry)
    #print(sequences)
    #print(names)
    return names,sequences

from openpyxl import Workbook
from openpyxl.styles import PatternFill

def write_to_excel(names, predictions, per_iden_list, output_filename="output.xlsx", 
                  mean_predictions=None, median_predictions=None, ci_lowers=None, 
                  ci_uppers=None, std_dev_list=None, hex_color_list=None, seq_lens_list=None):
    """
    Writes data to an Excel sheet, including bootstrap statistics and
    hexadecimal color codes, and colors the cells based on the hex codes.

    Args:
        names: List of names.
        predictions: List of predictions.
        per_iden_list: List of percentage identities.
        output_filename: Name of the output Excel file.
        mean_predictions: List of mean predictions (optional, for bootstrap).
        median_predictions: List of median predictions (optional, for bootstrap).
        ci_lowers: List of lower confidence intervals (optional, for bootstrap).
        ci_uppers: List of upper confidence intervals (optional, for bootstrap).
        std_dev_list: List of standard deviations (optional, for bootstrap).
        hex_color_list: List of hexadecimal color codes.
        seq_lens_list: List of sequence lengths
        """

    wb = Workbook()
    ws = wb.active

    
    if mean_predictions == None:
        ws.append(['Names', 'Single_Prediction', '%Identity_Nearest_VPOD_Sequence', 'Sequence_Length','Lmax_Hex_Color'])
        for i in range(len(names)):
            # Because openpyxel is picky about hex-codes we need to remove the '#' symbol for it to accept it as a fill color.
            hex_color = hex_color_list[i].replace('#','') 
            ws.append([names[i], predictions[i], per_iden_list[i], seq_lens_list[i], hex_color_list[i]])
            ws.cell(row=i+2, column=5).fill = PatternFill(start_color=hex_color, 
                                                        end_color=hex_color, 
                                                        fill_type="solid")
    else:
        ws.append(['Names', 'Single_Prediction', 'Prediction_Means', 'Prediction_Medians',
                    'Prediction_Lower_Bounds', 'Prediction_Upper_Bounds', 'Std_Deviation', 
                    '%Identity_Nearest_VPOD_Sequence', 'Sequence_Lengths','Lmax_Hex_Color'])
        for i in range(len(names)):
            # Because openpyxel is picky about hex-codes we need to remove the '#' symbol for it to accept it as a fill color.
            hex_color = hex_color_list[i].replace('#','') 
            ws.append([names[i], predictions[i], mean_predictions[i], median_predictions[i],
                        ci_lowers[i], ci_uppers[i], std_dev_list[i], per_iden_list[i], seq_lens_list[i], hex_color_list[i]])
            ws.cell(row=i+2, column=10).fill = PatternFill(start_color=hex_color, 
                                                        end_color=hex_color, 
                                                        fill_type="solid")
    wb.save(output_filename)
    
    